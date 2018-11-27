from openpyxl import load_workbook, Workbook
from classes import Team, Game, TeamManager, Level


def load_file(file, sheet="Sheet1"):

    team_manager = TeamManager()

    # load excel sheet into an iterable object
    wb = load_workbook(filename=file, read_only=True)
    ws = wb.active

    # loop through every row
    for row in ws.iter_rows(min_row=2):
        teamA = row[0].value    # Name of Team A
        scoreA = row[1].value   # Team A's Score

        scoreB = row[2].value   # Team B's Score
        teamB = row[3].value    # Name of Team B

        TeamA = team_manager.get_team(teamA)
        TeamB = team_manager.get_team(teamB)

        # if a team with the given name wasn't found in the list than create the team
        if not TeamA:
            TeamA = Team(teamA)

        if not TeamB:
            TeamB = Team(teamB)

        Match = Game(TeamA, TeamB, scoreA, scoreB)

        TeamA.add_game(Match)
        TeamB.add_game(Match)

        team_manager.add_team(TeamA)
        team_manager.add_team(TeamB)

    return team_manager

def write_file(teams, filename):
    wb = Workbook()
    ws = wb.active

    ws["A1"] = "Rank"
    ws["B1"] = "Team"
    ws["C1"] = "Score"

    rank = 1
    previous = Team("Empty")

    for i, team in enumerate(teams, start=2):
        rank_pos = "A" + str(i)
        team_pos = "B" + str(i)
        score_pos = "C" + str(i)

        if previous.points == team.points:
            ws[rank_pos] = rank - 1

        else:
            ws[rank_pos] = rank

        ws[team_pos] = team.get_name()
        ws[score_pos] = "%.2f" % team.get_points()

        rank += 1
        previous = team

    wb.save(filename)

