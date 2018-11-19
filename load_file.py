def load_file(file):
    from openpyxl import load_workbook
    from classes import Team, Game, Manager

    M = Manager()

    # load excel sheet into an iterable object
    wb = load_workbook(filename=file, read_only=True)
    ws = wb['Sheet1']

    # loop through every row
    for row in ws.iter_rows(min_row=2):
        teamA = row[0].value    # Name of Team A
        scoreA = row[1].value   # Team A's Score

        scoreB = row[2].value   # Team B's Score
        teamB = row[3].value    # Name of Team B

        TeamA = M.get_team(teamA)
        TeamB = M.get_team(teamB)

        # if Team with given name wasn't found in the list than create the team
        if not TeamA:
            TeamA = Team(teamA)

        if not TeamB:
            TeamB = Team(teamB)

        Match = Game(TeamA, TeamB, scoreA, scoreB)

        TeamA.add_game(Match)
        TeamB.add_game(Match)

        M.add_team(TeamA)
        M.add_team(TeamB)
        M.add_game(Match)

    return M

if __name__ == '__main__':
    Manager = load_file('data.xlsx')

    team = Manager.get_teams()[0]

    print(str(team) + "\n")

    for game in team.get_games():
        print(game)

    # games = Manager.get_games()
    #
    # for game in games:
    #     print(game)