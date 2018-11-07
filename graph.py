from load_file import load_teams
from openpyxl import load_workbook

class Graph(object):
    def __init__(self, file, alg="Win"):
        self.adj = {}   # Dictionary that maps a team to all of the nodes it is connected to.
        self.alg = alg

        wb = load_workbook(filename=file, read_only=True)
        ws = wb['Sheet1']
        # team_ids = load_teams(file) (I don't think we need this)
        # self.id_to_team_name = self.__get_id_to_team__(team_ids)

        for row in ws.iter_rows(min_row=2):
            teamA = row[0].value  # Name of Team A
            teamB = row[3].value  # Name of Team B

            scoreA = row[1].value  # Score of Team A
            scoreB = row[2].value  # Score of Team B

            self.__connect_nodes__(teamA, teamB, scoreA, scoreB)

        self.__calculate_relations__()

    def __get_id_to_team__(self, team_ids):
        id_to_team = {}
        for key, value in team_ids.items():
            id_to_team[value] = key

        return id_to_team

    def __connect_nodes__(self, nodeA, nodeB, scoreA, scoreB):

        """
        Structure of the Graph:

        dictionary of teams -->
            each team has a dictionary of other teams they faced -->
                other teams they faced have a list of scores that the other team got

        {
            teamA:  {
                        otherTeam:    [scores teamA got vs otherTeam] (list will be converted to ints after being processed),
                        anotherTeam:  [scores teamA got vs anotherTeam] (list will be converted to ints after being processed),
                        etc...
                    },

            teamB:  {
                        otherTeam:    [scores teamB got vs otherTeam] (list will be converted to ints after being processed),
                        anotherTeam:  [scores teamA got vs anotherTeam] (list will be converted to ints after being processed),
                        etc...
                    },
            etc...
        }
        """

        key_list = self.adj.keys() # list of all teams that are in the dictionary

        # if team already exists check if it has a connection to other team
        if nodeA in key_list:
            if nodeB in self.adj.get(nodeA).keys():
                self.adj.get(nodeA)[nodeB].append(scoreA)  # add team's score to the list of scores it got vs other team
            else:
                self.adj.get(nodeA)[nodeB] = [scoreA] # add team's score to the list of scores it got vs other team

        # if it doesn't exist add an entry with the team's relationship to the other team
        else:
            self.adj[nodeA] = {nodeB : [scoreA]}


        # same process but for other team
        if nodeB in key_list:
            if nodeA in self.adj.get(nodeB).keys():
                self.adj.get(nodeB)[nodeA].append(scoreB)
            else:
                self.adj.get(nodeB)[nodeA] = [scoreB]

        else:
            self.adj[nodeB] = {nodeA : [scoreB]}

    #  Prints the graph for testing purposes.
    def print_graph(self):
        print("%-20s%-25s%20s" % ("Team A", "Team B", "A's Score(s)"), end="")

        for key, value in self.adj.items():
            print("\n\n%-20s" % key)

            for key2, value2 in value.items():
                print("%-20s%-25s%20.2f" % ("", key2, value2))

    def __calculate_relations__(self):
        if self.alg == "Win":
            self.__win_alg__()

        elif self.alg == "Score":
            self.__score_alg__()

        elif self.alg == "Win Score":
            self.__win_score_alg__()

    # calculates each team's individual relation to the team they have played using win_alg
    def __win_alg__(self):
        pass

    def __score_alg__(self):
        pass

    def __win_score_alg__(self):
        pass

    def get_ranking_of(self, teamA):
        pass

graph = Graph("data.xlsx")
# graph.print_graph()

